"""
HR Data Validation Engine  –  OpenAI-Powered Edition
=====================================================
Compares two Excel files (Legacy/Paycor vs ADP) using SSN as the unique key.
OpenAI GPT-4o is used for:
  • Intelligent column mapping  (auto-detects column names in any format)
  • Discrepancy summary         (plain-English explanation of mismatches)
  • Missing-EE analysis         (flags patterns in missing employees)

Produces a 4-sheet validation workbook:
  1. Validation Sheet            – side-by-side Legacy | ADP | MATCH/ERROR/BLANK/MISMATCH
  2. Not in ADP and Legacy report – employees present in one system only
  3. Discrepancies               – field-level differences for matched employees
  4. Missing EE                  – employees absent from one system

Setup:
  1. pip install openai pandas openpyxl python-dotenv
  2. Create a .env file in this folder:
         OPENAI_API_KEY=sk-proj-xxxxxxxxxxxxxxxx
  3. Run:
         python validator_openai.py --legacy paycor.xlsx --adp adp.xlsx \
                                    --company "G&W Products" --output result.xlsx
"""

# ── Standard library ─────────────────────────────────────────────────────────
import argparse
import json
import re
import os
import sys
import warnings
from pathlib import Path

# ── Third-party ──────────────────────────────────────────────────────────────
try:
    from dotenv import load_dotenv
except ImportError:
    sys.exit("[ERROR] python-dotenv not installed.  Run:  pip install python-dotenv")

try:
    from openai import OpenAI
except ImportError:
    sys.exit("[ERROR] openai not installed.  Run:  pip install openai")

try:
    import pandas as pd
    import numpy as np
except ImportError:
    sys.exit("[ERROR] pandas not installed.  Run:  pip install pandas openpyxl")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

warnings.filterwarnings("ignore")

# ── Load .env ─────────────────────────────────────────────────────────────────
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
if not OPENAI_API_KEY:
    sys.exit(
        "[ERROR] OPENAI_API_KEY not found.\n"
        "  Create a .env file in this folder with:\n"
        "      OPENAI_API_KEY=sk-proj-xxxxxxxxxxxxxxxx\n"
        "  Get your key at: https://platform.openai.com/api-keys"
    )

import normalizer as norm

client = OpenAI(api_key=OPENAI_API_KEY, max_retries=0)
GPT_MODEL = "gpt-4o"          # change to "gpt-3.5-turbo" to reduce cost

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Fields are now imported from normalizer
REQUIRED_FIELDS   = norm.PERSONAL_FIELDS
PERSONAL_FIELDS   = norm.PERSONAL_FIELDS
JOB_FIELDS        = norm.JOB_FIELDS
TAX_FIELDS        = norm.TAX_FIELDS
COMPLIANCE_FIELDS = norm.COMPLIANCE_FIELDS
DEDUCTION_FIELDS  = norm.DEDUCTION_FIELDS
DIRECT_DEPOSIT_FIELDS = norm.DIRECT_DEPOSIT_FIELDS

# Core fields to always capture for metadata/context
CORE_FIELDS = ["Employment/Position Status", "Hire Date", "Termination Date", "Status"]

# Colours
GREEN_DARK  = "375623"
GREEN_LIGHT = "E2EFDA"
YELLOW_LIGHT = "FFFF99"
BLUE_LIGHT  = "DDEBF7"
RED_FILL    = "FF0000"
ORANGE_FILL = "FFA500"
BLANK_FILL  = "FFFF00"
MATCH_FILL  = "92D050"
WHITE       = "FFFFFF"
SALMON      = "FFC7CE"
DISC_HEADER = "C00000"
NAVY        = "1F4E79"
AI_FILL     = "EAF4FB"   # light blue – AI summary rows

# Global Meta Registry to capture "perfect" data from different sheets
METADATA_REGISTRY = {"legacy": {}, "adp": {}}

def _update_metadata_cache(df, cache, pk_col):
    """
    Iterate through normalized DataFrame and store Hire/Term dates + Status 
    for each SSN into the specified global cache.
    """
    for _, row in df.iterrows():
        pk = str(row.get(pk_col, "")).strip()
        if not pk or pk in ("nan", "None", ""): continue
        
        if pk not in cache: cache[pk] = {}
        
        # Capture canonical fields
        for f in ["Hire Date", "Termination Date", "Employment/Position Status", "Status"]:
            val = fmt_val(row.get(f, ""))
            # We store the "perfect" value (non-empty). 
            # If multiple sheets have it, we keep the first one found.
            if val and not cache[pk].get(f):
                cache[pk][f] = val

def _heal_from_metadata_cache(df, cache, pk_col):
    """
    Fill in missing Hire/Term dates + Status in the DataFrame using the global cache.
    """
    def heal_row(row):
        pk = str(row.get(pk_col, "")).strip()
        if not pk or pk not in cache: return row
        
        entry = cache[pk]
        for f in ["Hire Date", "Termination Date", "Employment/Position Status", "Status"]:
            current_val = fmt_val(row.get(f, ""))
            # If the field is currently missing or "Status" vs "Employment/Position Status" mismatch
            if not current_val and entry.get(f):
                row[f] = entry[f]
        return row
        
    return df.apply(heal_row, axis=1)
AI_FILL     = "EAF4FB"   # light blue – AI summary rows


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def hex_fill(h):
    return PatternFill("solid", start_color=h, fgColor=h)

def cell_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")

def thin_border():
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_cell(ws, row, col, value, bg=GREEN_DARK, fg=WHITE, bold=True):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = hex_fill(bg)
    c.font = cell_font(bold=bold, color=fg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
    return c

def _data_cell(ws, row, col, value, bg=WHITE, bold=False, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = hex_fill(bg)
    c.font = cell_font(bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = thin_border()
    return c

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def fmt_val(v):
    """Return empty string for blank/NaN, otherwise stripped string. 
       Strips time component (00:00:00) from dates.
    """
    if isinstance(v, pd.Series):
        v = v.dropna().iloc[0] if not v.dropna().empty else ""
    
    if pd.isna(v) or str(v).strip() in ("nan", "NaT", ""):
        return ""
    
    # Handle datetime objects
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    
    # Handle strings that look like dates with time (e.g. "2024-07-29 00:00:00")
    s = str(v).strip()
    if " 00:00:00" in s:
        s = s.replace(" 00:00:00", "")
    
    return s


# ---------------------------------------------------------------------------
# Step 1 – Load source files
# ---------------------------------------------------------------------------

def load_excel(path: str, sheet_idx: int = 0, id_col_hint: str = "SSN", deduplicate: bool = True) -> pd.DataFrame:
    path = Path(path)
    if not path.exists():
        raise ValueError(f"File not found: {path.name}")

    # Broad list of column names that represent an SSN/Tax ID
    SSN_ALIASES = {
        "ssn", "tax id", "taxid", "tax_id", "tax ssn", "ss_number",
        "social security", "social security number",
        "social security #", "ss#", "ss number",
        "emp ssn", "employee ssn", "tax id (ssn)", "tax id (ssn)#",
        "ee ssn", "payroll id", "individual id"
    }

    ACCOUNT_ALIASES = {
        "direct deposit account number", "account number", "account #", 
        "bank account number", "bank account #", "account", "account identifier"
    }

    # Metadata common fields to verify it's a real header (need at least 1-2)
    HEADER_MARKERS = {
        "name", "first name", "last name", "hire date", "birth date", 
        "gender", "sex", "job title", "department", "status",
        "deduction", "ded code", "deduction code", "ded description",
        "routing number", "account number", "net pay", "earnings"
    }

    with pd.ExcelFile(path) as xl:
        # Loop through sheets to find the correct one (prioritize sheet_idx if provided)
        target_sheets = [xl.sheet_names[sheet_idx]] + [s for i, s in enumerate(xl.sheet_names) if i != sheet_idx]
        
        found_sheet = None
        found_header_row = 0
        
        for sheet in target_sheets:
            # Scan up to 50 rows to find the header row
            try:
                raw = pd.read_excel(path, sheet_name=sheet, header=None, nrows=50)
            except Exception:
                continue

            for i, row in raw.iterrows():
                vals = [str(v).strip().lower() for v in row.values if pd.notna(v)]
                # Must have an ID alias AND at least one other marker to be sure it's the header
                id_aliases = ACCOUNT_ALIASES if "account" in id_col_hint.lower() else SSN_ALIASES
                
                # count_id check (exact or contains)
                count_id = 0
                for v in vals:
                    # Direct match or alias match
                    if v == id_col_hint.lower() or v in id_aliases:
                        count_id += 1
                    # Partial match for common variations if id_aliases didn't catch it
                    elif any(alias in v for alias in ["tax id", "ssn", "account #"]):
                        count_id += 1

                count_markers = sum(1 for v in vals if any(m in v for m in HEADER_MARKERS))
                if count_id >= 1 and count_markers >= 1:
                    found_sheet = sheet
                    found_header_row = i
                    break
            if found_sheet:
                break
        
        if not found_sheet:
            found_sheet = xl.sheet_names[0]
            found_header_row = 0

        df = pd.read_excel(path, sheet_name=found_sheet, header=found_header_row, dtype=str)

    df.columns = [str(c).strip() for c in df.columns]

    # Find the ID column (SSN or custom)
    target_id_col = id_col_hint
    id_col = next((c for c in df.columns if c.strip().lower() == target_id_col.lower()), None)
    
    # If not found exactly, try partial match or specific aliases
    if id_col is None:
        if "account" in target_id_col.lower():
            id_col = next((c for c in df.columns if c.strip().lower() in ACCOUNT_ALIASES), None)
        else:
            id_col = next((c for c in df.columns if c.strip().lower() in SSN_ALIASES), None)
    
    if id_col is None:
        # Fallback to partial match based on target hint
        search_terms = [target_id_col.lower()]
        if "account" in target_id_col.lower():
            search_terms.extend(["account", "bank"])
        else:
            search_terms.extend(["ssn", "tax id", "social"])
            
        id_col = next((c for c in df.columns if any(a in c.lower() for a in search_terms)), None)

    if id_col is None:
        raise ValueError(
            f"No ID column ('{target_id_col}') found in {path.name}. "
            f"Detected headers: {list(df.columns[:10])}..."
        )

    if id_col != target_id_col:
        df.rename(columns={id_col: target_id_col}, inplace=True)

    # Normalize primary key if it's SSN or Account
    if "SSN" in target_id_col:
        df[target_id_col] = df[target_id_col].apply(norm.normalize_ssn)
    elif "account" in target_id_col.lower():
        df[target_id_col] = df[target_id_col].apply(norm.normalize_account)
    
    # Filter out empty IDs
    df = df[df[target_id_col].astype(str).str.strip() != ""]
    df = df[~df[target_id_col].astype(str).isin(["nan", "NaT", "None"])]
    
    # Deduplicate by primary key
    if deduplicate:
        df = df.drop_duplicates(subset=target_id_col)
    return df


# ---------------------------------------------------------------------------
# Step 2 – OpenAI column mapping
# ---------------------------------------------------------------------------

def ai_map_columns(df: pd.DataFrame, source_label: str, required_fields: list) -> dict:
    """
    Ask GPT-4o to map the DataFrame's column names to required_fields.
    Returns { "Source Col": "Canonical Field" } for matched columns only.
    """
    cols = [c for c in df.columns if c not in ["Tax SSN", "SSN"]]
    prompt = f"""
You are an HR data expert. I have an Excel file from {source_label} with these column headers:
{json.dumps(cols, indent=2)}

Map each column to the best matching canonical field from this list:
{json.dumps(required_fields, indent=2)}

Rules:
- Only include columns you are confident about.
- Do NOT map "SSN" or "Tax ID (SSN)" (already handled).
- Return ONLY a JSON object like: {{"Source Col": "Canonical Field", ...}}
- If no good match exists for a column, omit it.
- Note: Deduction columns often have suffixes like "[All Deductions]" or symbols like "$" or "%". Map these to the base canonical field (e.g. "Deduction Code [All Deductions]" -> "Deduction Code").
"""
    print(f"   [AI] Mapping columns for {source_label} …")
    try:
        response = client.chat.completions.create(
            model=GPT_MODEL,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
            temperature=0,
            timeout=10,
        )
        mapping = json.loads(response.choices[0].message.content)
        print(f"   [AI] Mapped {len(mapping)} columns for {source_label}")
        return mapping
    except Exception as e:
        print(f"   [WARNING] AI Column Mapping for {source_label} failed: {e}")
        return {}


def apply_mapping(df: pd.DataFrame, mapping: dict, required_fields: list) -> pd.DataFrame:
    valid = {k: v for k, v in mapping.items() if k in df.columns}
    df = df.rename(columns=valid)
    return df


def ai_create_deduction_map(legacy_df: pd.DataFrame, adp_df: pd.DataFrame) -> dict:
    """
    Extract unique (Code, Description) from both sides and ask GPT-4o to map Legacy -> ADP.
    """
    if "Deduction Code" not in legacy_df.columns or "Deduction Code" not in adp_df.columns:
        return {}

    mapping = {}
    try:
        def get_pairs(df):
            cols = ["Deduction Code", "Deduction Description"]
            # Ensure they exist
            existing = [c for c in cols if c in df.columns]
            if not existing: return []
            subset = df[existing].drop_duplicates().dropna()
            # Convert everything to string to prevent JSON serialization errors
            subset = subset.astype(str)
            return subset.head(100).to_dict('records')

        leg_pairs = get_pairs(legacy_df)
        adp_pairs = get_pairs(adp_df)

        if not leg_pairs or not adp_pairs:
            return {}

        prompt = f"""
You are an HR data specialist. I need to map deduction categories between two payroll systems (Legacy and ADP).
Legacy uses descriptive codes, while ADP uses short abbreviations or IDs.

Legacy Unique Codes & Descriptions:
{json.dumps(leg_pairs, indent=2)}

ADP Unique Codes & Descriptions:
{json.dumps(adp_pairs, indent=2)}

Task: Create a mapping of Legacy "Deduction Code" to the corresponding ADP "Deduction Code".
Rules:
- Match based on descriptions (e.g., "Health Insurance" maps to "MEDICAL").
- If no match exists, do not include it.
- Return ONLY a JSON object: {{"LegacyCode": "ADPCode", ...}}
"""
        print(f"   [AI] Creating deduction code mapping …")
        response = client.chat.completions.create(
            model=GPT_MODEL,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
            temperature=0,
            timeout=10, # Shield against long hangs
        )
        mapping = json.loads(response.choices[0].message.content)
        print(f"   [AI] Mapped {len(mapping)} deduction codes")
    except Exception as e:
        print(f"   [WARNING] AI Deduction Mapping failed: {e}. Falling back to default matching.")
        mapping = {}
        
    return mapping


def apply_value_mapping(df: pd.DataFrame, col: str, mapping: dict) -> pd.DataFrame:
    """Apply a dictionary mapping to a specific column."""
    if col in df.columns:
        df[col] = df[col].map(lambda x: mapping.get(str(x), x))
    return df


# ---------------------------------------------------------------------------
# Step 3 – Comparison logic
# ---------------------------------------------------------------------------

def compare_values(v1, v2, field: str = ""):
    """Compare two values for the discrepancies sheet, with fuzzy logic for names."""
    b1 = pd.isna(v1) or str(v1).strip() in ("", "nan", "NaT")
    b2 = pd.isna(v2) or str(v2).strip() in ("", "nan", "NaT")
    if b1 and b2:
        return "MATCH"
    if b1 != b2:
        return "BLANK"
        
    s1, s2 = str(v1).strip(), str(v2).strip()
    if s1.lower() == s2.lower():
        return "MATCH"
    
    # Fuzzy name logic (handles Joel N. vs Joel Nicholas)
    is_name_field = field.lower().replace("_", " ") in ["ee name", "employee name", "name", "full name"]
    if is_name_field:
        # Normalize both and compare first/last. If they match and middle is consistent, it's a match.
        p1 = norm.normalize_full_name(s1)
        p2 = norm.normalize_full_name(s2)
        # p = (first, middle, last)
        f1, m1, l1 = [st.lower() for st in p1]
        f2, m2, l2 = [st.lower() for st in p2]
        
        if f1 == f2 and l1 == l2:
            # Middle name consistency check
            # Strip punctuation like dots for initial matching (e.g., N. vs Nicholas)
            m1_c = re.sub(r'[^a-z0-9]', '', m1)
            m2_c = re.sub(r'[^a-z0-9]', '', m2)
            
            if not m1_c or not m2_c: return "MATCH"
            if m1_c == m2_c: return "MATCH"
            if m1_c.startswith(m2_c) or m2_c.startswith(m1_c): return "MATCH"

    return "MISMATCH"

def comp_color(result):
    return {
        "MATCH":    MATCH_FILL,
        "MISMATCH": ORANGE_FILL,
        "ERROR":    RED_FILL,
        "BLANK":    BLANK_FILL,
    }.get(result, WHITE)


# ---------------------------------------------------------------------------
# Step 4 – OpenAI summaries
# ---------------------------------------------------------------------------

def ai_discrepancy_summary(records: list[dict]) -> str:
    """Generate a plain-English summary of discrepancy records."""
    if not records:
        return "No discrepancies found."
    try:
        sample = records[:40]   # keep tokens low
        prompt = f"""
You are an HR data analyst. Below are employee data discrepancies between a Legacy system and ADP.
Each record shows: Employee Name, field with mismatch, Legacy value, ADP value.

Discrepancies:
{json.dumps(sample, indent=2)}

Write a concise 3-5 sentence plain-English summary:
- How many employees are affected?
- Which fields have the most mismatches?
- What is the likely root cause?
- What action should HR take?
"""
        response = client.chat.completions.create(
            model=GPT_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3,
            max_tokens=300,
            timeout=10,
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"   [WARNING] AI Discrepancy Summary failed: {e}")
        return "AI Summary currently unavailable. Please check the Discrepancies sheet for details."


def ai_missing_ee_summary(only_legacy: list, only_adp: list) -> str:
    """Generate a plain-English summary of missing employees."""
    try:
        prompt = f"""
You are an HR data analyst. Compare two employee lists:

Employees in LEGACY but NOT in ADP ({len(only_legacy)} employees):
{json.dumps(only_legacy[:20], indent=2)}

Employees in ADP but NOT in LEGACY ({len(only_adp)} employees):
{json.dumps(only_adp[:20], indent=2)}

Write a concise 2-4 sentence summary:
- Are these terminated/retired employees that were cleaned up in one system?
- Any active employees missing that need urgent attention?
- What action should HR take?
"""
        response = client.chat.completions.create(
            model=GPT_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3,
            max_tokens=250,
            timeout=10,
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"   [WARNING] AI Missing EE Summary failed: {e}")
        return "AI Summary currently unavailable. Please check the Missing EE sheet for details."


# ---------------------------------------------------------------------------
# Step 5 – Build datasets
# ---------------------------------------------------------------------------

def build_datasets(legacy: pd.DataFrame, adp: pd.DataFrame, primary_key: str = "SSN"):
    id_col = primary_key
    
    if isinstance(id_col, list):
        # Handle composite keys (e.g., [SSN, Deduction Code])
        # Ensure all columns exist to avoid KeyError
        missing = [c for c in id_col if c not in legacy.columns or c not in adp.columns]
        if missing:
            raise ValueError(f"Primary key column(s) {missing} missing from mapped data. Check that your file contains deduction codes.")

        leg_set = set(map(tuple, legacy[id_col].values))
        adp_set = set(map(tuple, adp[id_col].values))
        matched  = leg_set & adp_set
        only_leg = leg_set - adp_set
        only_adp = adp_set - leg_set

        # Convert matched to a list of tuples for indexing
        matched_list = list(sorted(matched))
        
        # Use set_index to allow matching on the composite key
        ml = legacy.set_index(id_col, drop=False).loc[matched_list]
        ma = adp.set_index(id_col, drop=False).loc[matched_list]
    else:
        # Existing logic for single string key (SSN or Account Number)
        leg_ssns = set(legacy[id_col])
        adp_ssns = set(adp[id_col])
        matched  = leg_ssns & adp_ssns
        only_leg = leg_ssns - adp_ssns
        only_adp = adp_ssns - leg_ssns

        # Use loc instead of reindex to handle duplicate keys (like multiple deductions)
        ml = legacy[legacy[id_col].isin(matched)].set_index(id_col, drop=False).loc[sorted(matched)]
        ma = adp[adp[id_col].isin(matched)].set_index(id_col, drop=False).loc[sorted(matched)]
        
    return ml, ma, only_leg, only_adp


# ---------------------------------------------------------------------------
# Step 6 – Sheet builders
# ---------------------------------------------------------------------------

def build_validation_sheet(ws, ml, ma, company, required_fields, flat_header=False):
    n = len(required_fields)
    sp1 = n + 1                    # spacer col
    adp_s = n + 2
    sp2   = n + 1 + n + 1
    cmp_s = sp2 + 1

    if not flat_header:
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 45

        # Section headers row 1
        if "Gender" in required_fields:
            section_label = "PERSONAL"
            bg_l, fg_l, bg_a, fg_a = GREEN_DARK, WHITE, NAVY, WHITE
        elif "Do not Calculate F.U.T.A. Taxable?" in required_fields or "Federal/W4 Exemptions" in required_fields:
            section_label = "TAX"
            bg_l, fg_l, bg_a, fg_a = GREEN_DARK, WHITE, NAVY, WHITE
        elif "Ethnicity" in required_fields or "EEOC Job Classification" in required_fields:
            section_label = "COMPLIANCE"
            bg_l, fg_l, bg_a, fg_a = GREEN_DARK, WHITE, NAVY, WHITE
        elif "Account Number" in required_fields:
            section_label = "DIRECT DEPOSIT"
            bg_l, fg_l, bg_a, fg_a = GREEN_DARK, WHITE, NAVY, WHITE
        elif "Deduction Code" in required_fields:
            section_label = "DEDUCTION"
            # Use Blue for ADP (Right section in code, but ADP is usually Blue/Navy) 
            # and Peach for Legacy (Left section in code, but Legacy is usually Green)
            # Match Screenshot 2: ADP = Blue, Legacy = Peach
            bg_l, fg_l = "F8CBAD", "000000" # Legacy
            bg_a, fg_a = "DDEBF7", "000000" # ADP
        else:
            section_label = "JOB"
            bg_l, fg_l, bg_a, fg_a = GREEN_DARK, WHITE, NAVY, WHITE

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
        c = ws.cell(row=1, column=1, value=f"{section_label} INFO – LEGACY – {company}")
        c.fill = hex_fill(bg_l); c.font = cell_font(bold=True, color=fg_l, size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=1, column=sp1).fill = hex_fill("D9D9D9")

        ws.merge_cells(start_row=1, start_column=adp_s, end_row=1, end_column=adp_s + n - 1)
        c = ws.cell(row=1, column=adp_s, value=f"{section_label} INFO – ADP – {company}")
        c.fill = hex_fill(bg_a); c.font = cell_font(bold=True, color=fg_a, size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=1, column=sp2).fill = hex_fill("D9D9D9")

        ws.merge_cells(start_row=1, start_column=cmp_s, end_row=1, end_column=cmp_s + n - 1)
        c = ws.cell(row=1, column=cmp_s, value="COMPARISON RESULT")
        c.fill = hex_fill(BLUE_LIGHT); c.font = cell_font(bold=True, color="000000", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")

        header_row = 2
    else:
        header_row = 1
        bg_l, fg_l, bg_a, fg_a = GREEN_DARK, WHITE, NAVY, WHITE

    for i, f in enumerate(required_fields):
        _header_cell(ws, header_row, 1 + i,     f, bg=bg_l, fg=fg_l)
        _header_cell(ws, header_row, adp_s + i, f, bg=bg_a, fg=fg_a)
        _header_cell(ws, header_row, cmp_s + i, f, bg=BLUE_LIGHT, fg="000000")
    ws.cell(row=header_row, column=sp1).fill = hex_fill("D9D9D9")
    ws.cell(row=header_row, column=sp2).fill = hex_fill("D9D9D9")

    # Data rows
    data_start = header_row + 1
    for r, ssn in enumerate(ml.index, start=data_start):
        leg = ml.loc[ssn]
        adp = ma.loc[ssn]
        bg_l = GREEN_LIGHT if r % 2 else "C6EFCE"
        bg_a = YELLOW_LIGHT if r % 2 else "FFFACD"

        for i, field in enumerate(required_fields):
            vl = fmt_val(leg.get(field, ""))
            va = fmt_val(adp.get(field, ""))
            _data_cell(ws, r, 1 + i,     vl, bg=bg_l)
            _data_cell(ws, r, adp_s + i, va, bg=bg_a)

            # --- DYNAMIC FORMULA ---
            l_ref = f"{get_column_letter(1 + i)}{r}"
            a_ref = f"{get_column_letter(adp_s + i)}{r}"
            formula = (
                f'=IF((AND(ISBLANK({l_ref}),ISBLANK({a_ref}))),"BLANK",'
                f'(IF((NOT((ISBLANK({l_ref})=ISBLANK({a_ref})))),"ERROR",'
                f'IF({l_ref}={a_ref},"MATCH","MISMATCH"))))'
            )
            c = ws.cell(row=r, column=cmp_s + i, value=formula)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=r, column=sp1).fill = hex_fill("D9D9D9")
        ws.cell(row=r, column=sp2).fill = hex_fill("D9D9D9")

    # --- CONDITIONAL FORMATTING ---
    last_row = data_start + len(ml) - 1
    if last_row >= data_start:
        cmp_range = f"{get_column_letter(cmp_s)}{data_start}:{get_column_letter(cmp_s + n - 1)}{last_row}"
        ws.conditional_formatting.add(cmp_range, CellIsRule(operator='equal', formula=['"MATCH"'], fill=hex_fill(MATCH_FILL)))
        ws.conditional_formatting.add(cmp_range, CellIsRule(operator='equal', formula=['"MISMATCH"'], fill=hex_fill(ORANGE_FILL)))
        ws.conditional_formatting.add(cmp_range, CellIsRule(operator='equal', formula=['"ERROR"'], fill=hex_fill(RED_FILL)))
        ws.conditional_formatting.add(cmp_range, CellIsRule(operator='equal', formula=['"BLANK"'], fill=hex_fill(BLANK_FILL)))

    ws.freeze_panes = f"B{data_start}"
    widths = [18]*n + [2] + [18]*n + [2] + [14]*n
    _set_col_widths(ws, widths)


def build_not_in_report_sheet(ws, ml, ma, company, required_fields):
    # This sheet contains matched employees with a specific set of demographic/contact fields
    # If it's Personal info, use personal fields; Job -> job fields; Tax -> tax identifiers
    is_personal   = "Gender" in required_fields
    is_tax        = "Do not Calculate F.U.T.A. Taxable?" in required_fields or "Federal/W4 Exemptions" in required_fields
    is_dd         = "Account Number" in required_fields
    is_deduction  = "Deduction Code" in required_fields
    id_col        = "Account Number" if is_dd else "SSN"

    if is_personal:
        headers = ["EE Name", id_col, "Legal Middle Name", "Personal Email", 
                   "Work Email", "Cell/Mobile Phone", "Home Phone", "Work Phone", 
                   "Legal / Preferred Address: Address Line 2", "Marital Status"]
    elif is_tax:
        headers = ["EE Name", id_col, "Federal/W4 Marital Status Description",
                   "Federal/W4 Exemptions", "Federal Additional Tax Amount",
                   "Lived in State Code", "Worked in State Code",
                   "State Marital Status Code", "State Exemptions/Allowances",
                   "State Additional Tax Amount"]
    elif is_dd:
        headers = ["EE Name", id_col, "Direct Deposit Routing Number",
                   "Direct Deposit Amount", "Direct Deposit Amount Type",
                   "Direct Deposit Frequency", "SSN"]
    elif is_deduction:
        headers = ["EE Name", id_col, "Hire Date", "Termination Date",
                   "Deduction Code", "Deduction Description", "Deduction Rate"]
    else:
        headers = ["EE Name", id_col, "Job Title", "Home Department Description", 
                   "Business Unit Description", "Location Description", "Worker Category (FT, PT, TEMP, etc.)", 
                   "Reports To (Manager)", "FLSA (Exempt/Non Exempt)", "Annual Salary"]
    
    _header_cell(ws, 2, 2, "EE Name", bg=DISC_HEADER)
    for i, hdr in enumerate(headers[1:], start=3):
        _header_cell(ws, 2, i, hdr, bg=DISC_HEADER)

    for r, ssn in enumerate(ml.index, start=3):
        leg = ml.loc[ssn]
        name = _get_ee_name(leg)
        bg = "FFFFFF" if r % 2 == 0 else "F9F9F9"
        
        pk_val = headers[1] # Usually SSN or Account Number
        # If pk_val is a list (composite key), we just use the first element (SSN) for the display
        display_id = ssn[0] if isinstance(ssn, (list, tuple)) else ssn
        vals = [name, display_id]
        for h in headers[2:]:
            val = fmt_val(leg.get(h, ""))
            vals.append(val if val else "BLANK")
            
        for i, v in enumerate(vals, start=2):
            _data_cell(ws, r, i, v, bg=bg)
    _set_col_widths(ws, [3, 25, 16, 18, 25, 25, 25, 25, 25, 40, 15])


def build_discrepancies_sheet(ws, ml, ma, company, ai_summary: str, required_fields, id_col: str = "SSN"):
    # G&W headers: Client Name, SSN, EE Name, Status, Error in the Field, Legacy, ADP
    is_dd = "Account Number" in required_fields
    if is_dd:
        headers = ["Client Name", "SSN", "EE Name", "Status",
                   "Error in the Field", "Legacy", "ADP"]
    else:
        # If id_col is a list (composite), use the first element (SSN) as the header
        id_hdr = id_col[0] if isinstance(id_col, list) else id_col
        headers = ["Company Name", id_hdr, "EE Name", "Status",
                   "Error in the field", "Legacy", "ADP"]

    # AI Summary banner (row 2)
    ws.row_dimensions[2].height = 60
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=8)
    c = ws.cell(row=2, column=2, value=f"🤖 AI Summary: {ai_summary}")
    c.fill = hex_fill(AI_FILL)
    c.font = cell_font(bold=False, color="1F4E79", size=10)
    c.alignment = Alignment(wrap_text=True, vertical="center")

    current_row = 4
    disc_fields = [f for f in required_fields if f not in ["SSN", "Tax ID (SSN)", "Direct Deposit Account Number", "Account Number"]]

    # Build records grouped by field
    records_by_field: dict[str, list] = {f: [] for f in disc_fields}
    for pk_val in ml.index:
        leg = ml.loc[pk_val]
        adp = ma.loc[pk_val]
        name = _get_ee_name(leg)
        status = fmt_val(leg.get("Employment/Position Status", leg.get("Worker Category (FT, PT, TEMP, etc.)", ""))) or "Active"

        for field in disc_fields:
            vl = fmt_val(leg.get(field, ""))
            va = fmt_val(adp.get(field, ""))
            
            # Fallback for EE Name if it's blank in the field comparison
            if (field == "EE Name" or field == "Employee Name") and not va:
                va = _get_ee_name(adp)
                if va == "Unknown": va = ""

            res = compare_values(vl, va, field=field)
            if res != "MATCH":
                if is_dd:
                    records_by_field[field].append({
                        "Client Name": company,
                        "SSN": fmt_val(leg.get("EE SSN", pk_val)), 
                        "Status": status,
                        "EE Name": name,
                        "Error in the Field": field,
                        "Legacy": vl or "Data not available in legacy report",
                        "ADP":    va if va else "",
                        "Result": res,
                    })
                else:
                    id_hdr = id_col[0] if isinstance(id_col, list) else id_col
                    # Use the first element of pk_val (SSN) for the ID column
                    pk_display = pk_val[0] if isinstance(pk_val, (list, tuple)) else pk_val
                    records_by_field[field].append({
                        "Company Name": company,
                        id_hdr: pk_display,
                        "EE Name": name,
                        "Status": status,
                        "Error in the field": field,
                        "Legacy": vl or "Data not available in legacy report",
                        "ADP":    va if va else "",
                        "Result": res,
                    })

    for field in disc_fields:
        rows = records_by_field[field]
        
        # Section Label Row (Merged)
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=8)
        if not rows:
            c = ws.cell(row=current_row, column=2, value=f"✓ No discrepancies for: {field}")
            c.fill = hex_fill("E2EFDA"); c.font = cell_font(color="375623")
        else:
            c = ws.cell(row=current_row, column=2, value=f"Discrepancies for: {field}")
            c.fill = hex_fill("F8CBAD"); c.font = cell_font(color="C00000", bold=True)
        
        c.alignment = Alignment(horizontal="left")
        current_row += 1

        if rows:
            # Table Headers for this section
            for col, hdr in enumerate(headers, start=2):
                _header_cell(ws, current_row, col, hdr, bg=DISC_HEADER)
            current_row += 1

            # Data Rows
            for rec in rows:
                bg = SALMON if rec["Result"] == "MISMATCH" else "FFE0E0"
                for col, hdr in enumerate(headers, start=2):
                    _data_cell(ws, current_row, col, rec.get(hdr, ""), bg=bg)
                current_row += 1

        current_row += 1

    ws.freeze_panes = "B5"
    _set_col_widths(ws, [3, 18, 16, 22, 12, 30, 30, 30])


def build_missing_ee_sheet(ws, legacy, adp, only_leg, only_adp, ai_summary: str, id_col: str = "SSN"):
    is_dd = "Direct Deposit" in id_col or "Account" in id_col
    if is_dd:
        labels = ["EE SSN", "EE Name", "Status", "Hire Date", "Term Date"]
    else:
        # If id_col is list (composite), use the first one (SSN) as the header
        id_hdr = id_col[0] if isinstance(id_col, list) else id_col
        labels = [id_hdr, "Name", "Status", "Hire Date", "Term Date"]
    n = len(labels)
    sp = n + 2
    r_s = sp + 1

    # AI Summary
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=r_s + n - 1)
    c = ws.cell(row=2, column=2, value=f"🤖 AI Summary: {ai_summary}")
    c.fill = hex_fill(AI_FILL); c.font = cell_font(size=10); c.alignment = Alignment(wrap_text=True)

    # Left: ADP not in Legacy
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=n + 1)
    c = ws.cell(row=3, column=2, value="Employee available in ADP but not in LEGACY")
    c.fill = hex_fill(NAVY); c.font = cell_font(bold=True, color=WHITE)
    for i, lbl in enumerate(labels):
        _header_cell(ws, 4, 2 + i, lbl, bg=NAVY)

    if isinstance(id_col, list):
        mask = adp.set_index(id_col).index.isin(only_adp)
        adp_miss = adp[mask].reset_index(drop=True)
    else:
        adp_miss = adp[adp[id_col].isin(only_adp)].reset_index(drop=True)
    for r, row in adp_miss.iterrows():
        # Robust lookup for status and dates
        status = row.get("Status", row.get("Employment/Position Status", row.get("Worker Category (FT, PT, TEMP, etc.)", "")))
        h_date = row.get("Hire Date", "")
        t_date = row.get("Termination Date", "")
        
        # If id_col is list (composite), use the first one (SSN) for display
        display_id = row.get(id_col[0] if isinstance(id_col, list) else id_col, "")
        vals = [display_id, _get_ee_name(row), status, h_date, t_date]
        for i, v in enumerate(vals):
            _data_cell(ws, r + 5, 2 + i, fmt_val(v), bg=YELLOW_LIGHT if r%2 else "FFFACD")

    # Right: Legacy not in ADP
    ws.merge_cells(start_row=3, start_column=r_s, end_row=3, end_column=r_s + n - 1)
    c = ws.cell(row=3, column=r_s, value="Employee available in LEGACY but not in ADP")
    c.fill = hex_fill(GREEN_DARK); c.font = cell_font(bold=True, color=WHITE)
    for i, lbl in enumerate(labels):
        _header_cell(ws, 4, r_s + i, lbl, bg=GREEN_DARK)

    if isinstance(id_col, list):
        mask = legacy.set_index(id_col).index.isin(only_leg)
        leg_miss = legacy[mask].reset_index(drop=True)
    else:
        leg_miss = legacy[legacy[id_col].isin(only_leg)].reset_index(drop=True)
    for r, row in leg_miss.iterrows():
        # Robust lookup for status and dates
        status = row.get("Status", row.get("Employment/Position Status", row.get("Worker Category (FT, PT, TEMP, etc.)", "")))
        h_date = row.get("Hire Date", "")
        t_date = row.get("Termination Date", "")

        # If id_col is list (composite), use the first one (SSN) for display
        display_id = row.get(id_col[0] if isinstance(id_col, list) else id_col, "")
        vals = [display_id, _get_ee_name(row), status, h_date, t_date]
        for i, v in enumerate(vals):
            _data_cell(ws, r + 5, r_s + i, fmt_val(v), bg=GREEN_LIGHT if r%2 else "C6EFCE")

    for r in range(1, max(len(adp_miss), len(leg_miss)) + 5):
        ws.cell(row=r, column=sp).fill = hex_fill("D9D9D9")

    ws.freeze_panes = "B5"
    _set_col_widths(ws, [3] + [16]*n + [2] + [16]*n)

def _get_ee_name(df_row):
    fn = fmt_val(df_row.get("Legal First Name", ""))
    ln = fmt_val(df_row.get("Legal Last Name", ""))
    if fn and ln: return f"{ln}, {fn}"
    full = fmt_val(df_row.get("EE Name", df_row.get("Employee Full Name", "")))
    if full: return full
    return fn or ln or fmt_val(df_row.get("SSN", df_row.get("Tax ID (SSN)", "")))


def build_code_mapping_sheet(ws, legacy_df: pd.DataFrame, adp_df: pd.DataFrame):
    """
    Build a 'Code Mapping' sheet that lists unique Deduction Code → Description pairs
    sourced from both Legacy and ADP dataframes in a side-by-side layout.
    """
    BLUE_LIGHT = "DDEBF7"  # ADP sections
    PEACH = "F8CBAD"       # Legacy sections
    GAP_COL = "D9D9D9"     # Spacer color

    # Clear Row 1 if needed (ensure it's clean)
    ws.row_dimensions[1].height = 15

    # Header Row 2: Merged labels for ADP and Legacy
    ws.row_dimensions[2].height = 25
    
    # ADP Header
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=3)
    c_adp = ws.cell(row=2, column=2, value="ADP")
    c_adp.fill = hex_fill(BLUE_LIGHT)
    c_adp.font = cell_font(bold=True, size=11)
    c_adp.alignment = Alignment(horizontal="center", vertical="center")
    c_adp.border = thin_border()
    ws.cell(row=2, column=3).border = thin_border()

    # Legacy Header
    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=6)
    c_leg = ws.cell(row=2, column=5, value="Legacy")
    c_leg.fill = hex_fill(PEACH)
    c_leg.font = cell_font(bold=True, size=11)
    c_leg.alignment = Alignment(horizontal="center", vertical="center")
    c_leg.border = thin_border()
    ws.cell(row=2, column=6).border = thin_border()

    # Header Row 3: Column field names
    _header_cell(ws, 3, 2, "Deduction Code", bg=BLUE_LIGHT, fg="000000")
    _header_cell(ws, 3, 3, "Deduction Description", bg=BLUE_LIGHT, fg="000000")
    
    _header_cell(ws, 3, 5, "Deduction Code", bg=PEACH, fg="000000")
    _header_cell(ws, 3, 6, "Deduction Description", bg=PEACH, fg="000000")

    # Helper to get unique (Code, Description) pairs
    def get_unique_pairs(df):
        if df is None or df.empty: return []
        
        # Robust column identification
        code_col = next((c for c in df.columns if c.lower() in norm.DED_CODE_FIELDS), None)
        desc_col = next((c for c in df.columns if c.lower() in norm.DED_DESC_FIELDS), None)
        
        if not code_col: return []
        
        pairs = {}
        for _, row in df.iterrows():
            code = fmt_val(row.get(code_col, ""))
            desc = fmt_val(row.get(desc_col, "")) if desc_col else ""
            if code and code not in pairs:
                pairs[code] = desc
        return sorted(pairs.items())

    adp_pairs = get_unique_pairs(adp_df)
    leg_pairs = get_unique_pairs(legacy_df)

    # Write Data starting from Row 4
    max_len = max(len(adp_pairs), len(leg_pairs))
    for i in range(max_len):
        row_idx = 4 + i
        bg = WHITE
        
        # ADP Data (Cols B & C)
        if i < len(adp_pairs):
            code, desc = adp_pairs[i]
            _data_cell(ws, row_idx, 2, code, bg=bg)
            _data_cell(ws, row_idx, 3, desc, bg=bg)
        else:
            _data_cell(ws, row_idx, 2, "", bg=bg)
            _data_cell(ws, row_idx, 3, "", bg=bg)

        # Gap Column D (Col 4)
        ws.cell(row=row_idx, column=4).fill = hex_fill(WHITE) # Keep gap clean
        
        # Legacy Data (Cols E & F)
        if i < len(leg_pairs):
            code, desc = leg_pairs[i]
            _data_cell(ws, row_idx, 5, code, bg=bg)
            _data_cell(ws, row_idx, 6, desc, bg=bg)
        else:
            _data_cell(ws, row_idx, 5, "", bg=bg)
            _data_cell(ws, row_idx, 6, "", bg=bg)

    # Styling and widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 45
    
    ws.freeze_panes = "A4"

# ---------------------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------------------

def run_validation(legacy_path, adp_path, company, output_path, required_fields, primary_key: str = "SSN", sheet_idx: int = 0):

    print(f"\n[Validation] Processing: {output_path}")

    print(f"Phase 1: AI column mapping (Legacy, Sheet {sheet_idx}) …")
    legacy_raw = load_excel(legacy_path, sheet_idx=sheet_idx, id_col_hint=primary_key)
    norm_fields = list(dict.fromkeys(required_fields + CORE_FIELDS)) # Ensure uniqueness
    leg_map = ai_map_columns(legacy_raw, "Legacy/Paycor", norm_fields)
    legacy  = apply_mapping(legacy_raw.copy(), leg_map, norm_fields)
    legacy  = norm.normalize_dataframe(legacy, norm_fields)

    print(f"\nPhase 2: AI column mapping (ADP, Sheet {sheet_idx}) …")
    adp_raw = load_excel(adp_path, sheet_idx=sheet_idx, id_col_hint=primary_key)
    adp_map = ai_map_columns(adp_raw, "ADP", norm_fields)
    adp     = apply_mapping(adp_raw.copy(), adp_map, norm_fields)
    adp     = norm.normalize_dataframe(adp, norm_fields)

    # Phase 2.5: Global Metadata Healing & Capture
    # (Capture values from this sheet for other sheets, and fill holes from previous sheets)
    _update_metadata_cache(legacy, METADATA_REGISTRY["legacy"], primary_key)
    _update_metadata_cache(adp,    METADATA_REGISTRY["adp"],    primary_key)
    
    legacy = _heal_from_metadata_cache(legacy, METADATA_REGISTRY["legacy"], primary_key)
    adp    = _heal_from_metadata_cache(adp,    METADATA_REGISTRY["adp"],    primary_key)

    print(f"\nPhase 3: Matching employees by {primary_key} …")
    ml, ma, only_leg, only_adp = build_datasets(legacy, adp, primary_key=primary_key)
    
    # FALLBACK 1: If Direct Deposit on sheet 1 yielded no matches, try sheet 0
    if len(ml) == 0 and "account" in primary_key.lower() and sheet_idx == 1:
        print("      [Fallback] No matches on sheet 1 for Direct Deposit. Trying sheet 0 …")
        legacy_raw = load_excel(legacy_path, sheet_idx=0, id_col_hint=primary_key)
        legacy = apply_mapping(legacy_raw.copy(), leg_map, required_fields)
        legacy = norm.normalize_dataframe(legacy, required_fields)
        
        adp_raw = load_excel(adp_path, sheet_idx=0, id_col_hint=primary_key)
        adp = apply_mapping(adp_raw.copy(), adp_map, required_fields)
        adp = norm.normalize_dataframe(adp, required_fields)
        
        ml, ma, only_leg, only_adp = build_datasets(legacy, adp, primary_key=primary_key)

    # FALLBACK 2: If Deduction on sheet 0 yielded no matches, try searching for "Deduction" sheets or sheet index 2
    is_deduction = "Deduction Code" in required_fields
    if len(ml) == 0 and is_deduction and sheet_idx == 0:
        print("      [Fallback] No matches on sheet 0 for Deduction. Searching for 'Deduction' sheets …")
        # Try sheet index 2 first (3rd sheet)
        for try_idx in [2, 1]:
            try:
                legacy_raw = load_excel(legacy_path, sheet_idx=try_idx, id_col_hint=primary_key)
                adp_raw    = load_excel(adp_path,    sheet_idx=try_idx, id_col_hint=primary_key)
                
                # Re-map and re-normalize
                l_map = ai_map_columns(legacy_raw, f"Legacy (Sheet {try_idx})", norm_fields)
                l_df  = apply_mapping(legacy_raw.copy(), l_map, norm_fields)
                l_df  = norm.normalize_dataframe(l_df, norm_fields)
                
                a_map = ai_map_columns(adp_raw, f"ADP (Sheet {try_idx})", norm_fields)
                a_df  = apply_mapping(adp_raw.copy(), a_map, norm_fields)
                a_df  = norm.normalize_dataframe(a_df, norm_fields)
                
                ml_t, ma_t, ol_t, oa_t = build_datasets(l_df, a_df, primary_key=primary_key)
                if len(ml_t) > 0:
                    print(f"      [Success] Found {len(ml_t)} matches on sheet {try_idx}")
                    ml, ma, only_leg, only_adp = ml_t, ma_t, ol_t, oa_t
                    legacy, adp = l_df, a_df
                    break
            except Exception:
                continue

    print(f"      Matched: {len(ml)}  |  Only Legacy: {len(only_leg)}  |  Only ADP: {len(only_adp)}\n")

    print("Phase 4: AI discrepancy summary …")
    # Build flat list for AI summary
    disc_sample = []
    # Identify name fields correctly
    f_name = "Legal First Name" if "Legal First Name" in required_fields else "Employee Full Name"
    l_name = "Legal Last Name"
    
    for pk_val in ml.index:
        leg = ml.loc[pk_val]; adp_row = ma.loc[pk_val]
        for field in required_fields:
            if field in ["SSN", "Tax ID (SSN)", "Direct Deposit Account Number", "Account Number", "EE SSN"]: continue
            vl = fmt_val(leg.get(field, ""))
            va = fmt_val(adp_row.get(field, ""))
            if compare_values(vl, va) not in ("MATCH", "BLANK"):
                name = _get_ee_name(leg)
                disc_sample.append({
                    "Employee": name,
                    "Field": field, "Legacy": vl, "ADP": va
                })
    disc_ai = ai_discrepancy_summary(disc_sample)

    print("Phase 5: AI missing-EE summary …")
    leg_id = primary_key
    adp_id = primary_key
    
    leg_names = []
    for s in list(only_leg)[:20]:
        rows = legacy[legacy[leg_id] == s]
        if not rows.empty:
            leg_names.append(_get_ee_name(rows.iloc[0]))
        else:
            leg_names.append(str(s))

    adp_names = []
    for s in list(only_adp)[:20]:
        rows = adp[adp[adp_id] == s]
        if not rows.empty:
            adp_names.append(_get_ee_name(rows.iloc[0]))
        else:
            adp_names.append(str(s))
            
    miss_ai = ai_missing_ee_summary(leg_names, adp_names)

    print("Phase 6: Writing workbook …")
    wb = Workbook()

    ws1 = wb.active
    is_dd         = "Direct Deposit Account Number" in required_fields
    is_deduction  = "Deduction Code" in required_fields
    
    if is_deduction:
        ws1.title = "Deduction Validation"
        build_validation_sheet(ws1, ml, ma, company, required_fields, flat_header=False)
    else:
        ws1.title = "Validation Sheet" if is_dd else "Validated Data"
        build_validation_sheet(ws1, ml, ma, company, required_fields)

    ws2 = wb.create_sheet("Discrepancies")
    id_col_label = "Account Number" if primary_key == "Account Number" else "SSN"
    build_discrepancies_sheet(ws2, ml, ma, company, disc_ai, required_fields, id_col=id_col_label)

    ws3 = wb.create_sheet("Missing EE")
    build_missing_ee_sheet(ws3, legacy, adp, only_leg, only_adp, miss_ai, id_col=id_col_label)

    if not is_dd and not is_deduction:
        # Sheet 4 name varies by field group
        if "Gender" in required_fields:
            sheet4_name = "Not in Legacy , not in adp"
        else:
            sheet4_name = "Not in Legacy Not in adp"
        ws4 = wb.create_sheet(sheet4_name)
        build_not_in_report_sheet(ws4, ml, ma, company, required_fields)

    # Prepare detailed data for the UI
    validation_sheet_data = []
    for pk_val in ml.index:
        leg_row = ml.loc[pk_val]
        adp_row = ma.loc[pk_val]
        
        # Handle duplicates: ensure we are working with a single Series (row)
        if hasattr(leg_row, "iloc") and len(leg_row.shape) > 1: leg_row = leg_row.iloc[0]
        if hasattr(adp_row, "iloc") and len(adp_row.shape) > 1: adp_row = adp_row.iloc[0]
        
        name = _get_ee_name(leg_row)
        
        fields = {}
        for field in required_fields:
            vl = fmt_val(leg_row.get(field, ""))
            va = fmt_val(adp_row.get(field, ""))
            
            # Simple Python version of the Excel formula
            if not vl and not va:
                status = "BLANK"
            elif (not vl and va) or (vl and not va):
                status = "ERROR"
            elif str(vl).strip() == str(va).strip():
                status = "MATCH"
            else:
                status = "MISMATCH"
                
            fields[field] = {
                "legacy": vl,
                "adp": va,
                "status": status
            }
        
        # If pk_val is composite (list/tuple), use first element for display
        display_id = pk_val[0] if isinstance(pk_val, (list, tuple)) else pk_val
        
        validation_sheet_data.append({
            "id": str(display_id),
            "employeeName": name,
            "employeeId": str(display_id),
            "fields": fields
        })

    # Build Discrepancy list for UI
    discrepancy_data = []
    for item in disc_sample:
        # Use a flexible ID lookup to avoid KeyError
        eid = item.get("SSN", item.get("Account Number", item.get("ID", "Unknown")))
        discrepancy_data.append({
            "id": f"{eid}_{item.get('Field', 'unknown')}",
            "employeeName": item.get("Name", "Unknown"),
            "employeeId": str(eid),
            "field": item.get("Field", "unknown"),
            "legacyValue": str(item.get("Legacy", "")),
            "adpValue": str(item.get("ADP", ""))
        })

    # Build Missing lists
    missing_adp_data = []
    for ssn in only_leg:
        eid = str(ssn)
        # Find row in legacy df for name - safer filtering
        mask = legacy[primary_key].astype(str) == eid
        rows = legacy[mask]
        name = _get_ee_name(rows.iloc[0]) if not rows.empty else "Unknown"
        missing_adp_data.append({
            "id": eid,
            "employeeName": name,
            "employeeId": eid,
            "source": "legacy"
        })

    missing_legacy_data = []
    for ssn in only_adp:
        eid = str(ssn)
        # Find row in adp df for name - safer filtering
        mask = adp[primary_key].astype(str) == eid
        rows = adp[mask]
        name = _get_ee_name(rows.iloc[0]) if not rows.empty else "Unknown"
        missing_legacy_data.append({
            "id": eid,
            "employeeName": name,
            "employeeId": eid,
            "source": "adp"
        })

    wb.save(output_path)
    print(f"\n[OK]  Saved -> {output_path}\n")

    return {
        "summary": {
            "totalEmployees": len(ml) + len(only_leg) + len(only_adp),
            "matched": len(ml),
            "mismatched": len(disc_sample),
            "missingInADP": len(only_leg),
            "missingInLegacy": len(only_adp)
        },
        "validationSheet": validation_sheet_data,
        "discrepancies": discrepancy_data,
        "missingInADP": missing_adp_data,
        "missingInLegacy": missing_legacy_data
    }

def load_deduction_mapping_strict(path: str) -> pd.DataFrame:
    """Load Sheet 4 (index 3) and find the headers for mapping."""
    try:
        xl = pd.ExcelFile(path)
        # Filter sheet names to find one with 'mapping' or index 3
        sheet_name = None
        for s in xl.sheet_names:
            if "mapping" in s.lower() or "sheet 4" in s.lower():
                sheet_name = s
                break
        if not sheet_name:
            sheet_name = xl.sheet_names[3] if len(xl.sheet_names) > 3 else xl.sheet_names[0]
        
        # Read first 10 rows to find header row containing Code_ID or Deduction Code
        raw = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=10)
        header_row = 1 
        for i, row in raw.iterrows():
            vals = [str(v).lower() for v in row.values if pd.notna(v)]
            if "code_id" in vals or "deduction code" in vals or "common code" in vals:
                header_row = i
                break
                
        mapping_df = pd.read_excel(path, sheet_name=sheet_name, header=header_row)
        mapping_df.columns = [str(c).strip() for c in mapping_df.columns]
        return mapping_df
    except:
        return pd.DataFrame()

def run_deduction_validation(legacy_path, adp_path, company, output_path):
    print(f"\n[Deduction] Processing: {output_path}")
    
    # 1. Load Mapping Sheets from BOTH files
    leg_map_df = load_deduction_mapping_strict(legacy_path)
    adp_map_df = load_deduction_mapping_strict(adp_path)
    
    print(f"Phase 1: Loading Deduction Mapping from Sheet 4 …")
    
    def get_id_map(m_df):
        if m_df.empty: return {}, {}
        m_cols = m_df.columns
        id_col = next((c for c in m_cols if "code_id" in c.lower() or "common code" in c.lower()), None)
        code_col = next((c for c in m_cols if "deduction code" in c.lower()), None)
        desc_col = next((c for c in m_cols if "description" in c.lower()), None)
        
        if not id_col or not code_col: return {}, {}
        
        id_map = m_df[[code_col, id_col]].dropna().astype(str).set_index(code_col)[id_col].to_dict()
        desc_map = m_df[[id_col, desc_col]].dropna().astype(str).set_index(id_col)[desc_col].to_dict() if desc_col else {}
        return id_map, desc_map

    leg_to_id, leg_id_to_desc = get_id_map(leg_map_df)
    adp_to_id, adp_id_to_desc = get_id_map(adp_map_df)
    
    # Merge mappings if possible (ID to Desc)
    master_id_to_desc = {**leg_id_to_desc, **adp_id_to_desc}

    # 2. Extract Deduction Data
    req_fields = norm.DEDUCTION_FIELDS # ['SSN', 'Full Name', 'Deduction Code', 'Deduction Description', 'Deduction Amount', 'Deduction Rate']
    
    def get_ded_df(path, label):
        xl = pd.ExcelFile(path)
        sheet = next((s for s in xl.sheet_names if "deduction" in s.lower()), xl.sheet_names[0])
        # IMPORTANT: deduplicate=False to keep all deductions for the same employee
        df_raw = load_excel(path, sheet_idx=xl.sheet_names.index(sheet), id_col_hint="SSN", deduplicate=False)
        mapping = ai_map_columns(df_raw, label, req_fields)
        df = apply_mapping(df_raw.copy(), mapping, req_fields)
        df = norm.normalize_dataframe(df, req_fields)
        
        # Unify name to "Full Name" for matching
        if "Employee Full Name" in df.columns:
            df["Full Name"] = df["Employee Full Name"]
        elif "Legal First Name" in df.columns and "Legal Last Name" in df.columns:
            df["Full Name"] = df.apply(lambda r: f"{r['Legal Last Name']}, {r['Legal First Name']}".upper().strip(), axis=1)
        
        return df

    legacy_ded = get_ded_df(legacy_path, "Legacy Deduction")
    adp_ded = get_ded_df(adp_path, "ADP Deduction")
    
    # 3. Apply Code_ID mapping
    legacy_ded["Code_ID"] = legacy_ded["Deduction Code"].astype(str).map(leg_to_id).apply(lambda x: str(x).replace(".0", "") if pd.notna(x) else "")
    adp_ded["Code_ID"] = adp_ded["Deduction Code"].astype(str).map(adp_to_id).apply(lambda x: str(x).replace(".0", "") if pd.notna(x) else "")
    
    # Clean up names for matching (Extremely Robust: Ignore middle names, initials, suffixes, and nicknames)
    def make_robust_key(row):
        fname = str(row.get("Full Name", "")).strip()
        first, middle, last = norm.normalize_full_name(fname)
        
        # 1. Standardize and remove suffixes from last name
        last = last.upper()
        for suffix in [" JR", " SR", " II", " III", " IV", " V"]:
            if last.endswith(suffix):
                last = last.replace(suffix, "").strip()
        
        # 2. Use only first 3 letters of first name to handle nicknames (Ken/Kenneth)
        first_short = first.upper()[:3]
        
        # 3. Final clean: remove all non-alpha
        clean_f = re.sub(r'[^A-Z]', '', first_short)
        clean_l = re.sub(r'[^A-Z]', '', last)
        
        cid = str(row.get("Code_ID", "")).strip()
        return f"{clean_l}_{clean_f}_{cid}"

    legacy_ded["MatchName"] = legacy_ded.apply(make_robust_key, axis=1)
    adp_ded["MatchName"] = adp_ded.apply(make_robust_key, axis=1)
    
    # --- New Robust Alignment Logic ---
    # To handle multiple deductions per Code_ID and ensure 100% accuracy:
    # 1. We create an 'InstanceID' to distinguish multiple deductions of the same Code_ID for one person
    legacy_ded["InstanceID"] = legacy_ded.groupby(["MatchName", "Code_ID"]).cumcount()
    adp_ded["InstanceID"] = adp_ded.groupby(["MatchName", "Code_ID"]).cumcount()
    
    # 2. Use (MatchName, Code_ID, InstanceID) as the primary key for matching
    pk = ["MatchName", "Code_ID", "InstanceID"]
    ml, ma, only_leg, only_adp = build_datasets(legacy_ded, adp_ded, primary_key=pk)
    
    print(f"      Matched: {len(ml)}  |  Only Legacy: {len(only_leg)}  |  Only ADP: {len(only_adp)}")

    # 4. Generate Report
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Validated Data"
    v_fields = ["SSN", "Full Name", "Code_ID", "Deduction Code", "Deduction Description", "Deduction Amount", "Deduction Rate"]
    build_validation_sheet(ws1, ml, ma, company, v_fields, flat_header=False)
    
    # Discrepancies
    ws2 = wb.create_sheet("Discrepancies")
    disc_sample = []
    # Now that we aligned by InstanceID, ml and ma have the same length and 1-to-1 rows!
    for i in range(len(ml)):
        l_row, a_row = ml.iloc[i], ma.iloc[i]
        for f in ["Deduction Description", "Deduction Amount", "Deduction Rate"]:
            vl, va = fmt_val(l_row.get(f, "")), fmt_val(a_row.get(f, ""))
            if compare_values(vl, va) != "MATCH":
                disc_sample.append({"Employee": l_row["Full Name"], "Field": f, "Legacy": vl, "ADP": va})

    disc_ai = ai_discrepancy_summary(disc_sample)
    build_discrepancies_sheet(ws2, ml, ma, company, disc_ai, v_fields, id_col=pk)
    
    # Missing Employee
    ws3 = wb.create_sheet("Missing Employee")
    l_names = [legacy_ded[legacy_ded.set_index(pk).index == s].iloc[0]["Full Name"] for s in list(only_leg)[:15]]
    a_names = [adp_ded[adp_ded.set_index(pk).index == s].iloc[0]["Full Name"] for s in list(only_adp)[:15]]
    miss_ai = ai_missing_ee_summary(l_names, a_names)
    build_missing_ee_sheet(ws3, legacy_ded, adp_ded, only_leg, only_adp, miss_ai, id_col=pk)
    
    # Not in ADP
    ws4 = wb.create_sheet("Not in ADP")
    build_not_in_report_sheet(ws4, ml, ma, company, v_fields)
    
    # Code Mapping (Combined)
    ws5 = wb.create_sheet("Code Mapping")
    m_combined = pd.concat([leg_map_df, adp_map_df]).drop_duplicates().sort_values(by=leg_map_df.columns[0] if not leg_map_df.empty else 0)
    for ci, col in enumerate(m_combined.columns, 1):
        _header_cell(ws5, 1, ci, col, bg="DDEBF7", fg="000000")
        for ri, val in enumerate(m_combined[col], 2):
            _data_cell(ws5, ri, ci, fmt_val(val))

    wb.save(output_path)
    print(f"[OK] Saved -> {output_path}")

    return {
        "summary": {"totalEmployees": len(ml) + len(only_leg) + len(only_adp), "matched": len(ml), "mismatched": len(disc_sample), "missingInADP": len(only_leg), "missingInLegacy": len(only_adp)},
        "validationSheet": [], "discrepancies": [], "missingInADP": [], "missingInLegacy": []
    }



# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="HR Data Validation Engine – OpenAI-powered"
    )
    parser.add_argument("--legacy",  required=True,
                        help="Path to Legacy/Paycor Excel file")
    parser.add_argument("--adp",     required=True,
                        help="Path to ADP Excel file")
    parser.add_argument("--company", default="Company",
                        help="Company name label  (default: Company)")
    parser.add_argument("--output",  default="validation_output.xlsx",
                        help="Output Excel file path")
    args = parser.parse_args()

    # Run Personal Info Validation
    pers_output = args.output.replace(".xlsx", " - Personal Info Validation.xlsx")
    if args.output == "validation_output.xlsx":
         pers_output = f"{args.company} - Personal Info Validation (Output).xlsx"

    run_validation(
        legacy_path = args.legacy,
        adp_path    = args.adp,
        company     = args.company,
        output_path = pers_output,
        required_fields = PERSONAL_FIELDS
    )

    # Run Direct Deposit Validation
    dd_output = args.output.replace(".xlsx", " - Direct Deposit Info Validation.xlsx")
    if args.output == "validation_output.xlsx":
         dd_output = f"{args.company} - Direct Deposit Validation (Output).xlsx"

    run_validation(
        legacy_path = args.legacy,
        adp_path    = args.adp,
        company     = args.company,
        output_path = dd_output,
        required_fields = norm.DIRECT_DEPOSIT_FIELDS,
        primary_key = "Account Number",
        sheet_idx = 1
    )

    # Run Job Info Validation
    job_output = args.output.replace(".xlsx", " - Job Info Validation.xlsx")
    if args.output == "validation_output.xlsx":
         job_output = f"{args.company} - Job Information Validation (Output).xlsx"

    run_validation(
        legacy_path = args.legacy,
        adp_path    = args.adp,
        company     = args.company,
        output_path = job_output,
        required_fields = JOB_FIELDS
    )

    # Run Tax Info Validation
    tax_output = args.output.replace(".xlsx", " - Tax Info Validation.xlsx")
    if args.output == "validation_output.xlsx":
         tax_output = f"{args.company} - Tax Information Validation (Output).xlsx"

    run_validation(
        legacy_path = args.legacy,
        adp_path    = args.adp,
        company     = args.company,
        output_path = tax_output,
        required_fields = TAX_FIELDS
    )

    # Run Compliance Info Validation
    comp_output = args.output.replace(".xlsx", " - Compliance Info Validation.xlsx")
    if args.output == "validation_output.xlsx":
         comp_output = f"{args.company} - Compliance Information Validation (Output).xlsx"

    run_validation(
        legacy_path = args.legacy,
        adp_path    = args.adp,
        company     = args.company,
        output_path = comp_output,
        required_fields = COMPLIANCE_FIELDS
    )

    # Run Deduction Info Validation
    ded_output = args.output.replace(".xlsx", " - Deduction Info Validation.xlsx")
    if args.output == "validation_output.xlsx":
         ded_output = f"{args.company} - Deduction Information Validation (Output).xlsx"

    run_deduction_validation(
        legacy_path = args.legacy,
        adp_path    = args.adp,
        company     = args.company,
        output_path = ded_output
    )


if __name__ == "__main__":
    main()
